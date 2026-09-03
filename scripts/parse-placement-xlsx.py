#!/usr/bin/env python3
"""
parse-placement-xlsx.py
=======================

Single-pass parser that converts the TPO-maintained "ALL PLACED STUDENT
LIST" XLSX into the structured JSON the website reads at build time.

This is the ONE conversion step in the "placement data design system":

    data/source/all-placed-students.xlsx   (TPO source of truth)
                  ↓ [this script]
    lib/alumni-manifest.json               (structured manifest)
                  ↓ [lib/placement-stats.ts]
    every public surface that mentions placement numbers

To update placement numbers everywhere on the site, replace the XLSX
in data/source/, re-run this script, and commit. No need to grep for
"1,331" across the codebase — the constants in lib/placement-stats.ts
are derived from the manifest and cascade automatically.

Usage:
    python3 scripts/parse-placement-xlsx.py

Outputs:
    lib/alumni-manifest.json   (the structured manifest)

Requires:
    pip install openpyxl
"""

from __future__ import annotations

import collections
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Install openpyxl first:  pip3 install openpyxl")

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT / "data" / "source" / "all-placed-students.xlsx"
OUTPUT_MANIFEST = REPO_ROOT / "lib" / "alumni-manifest.json"

DATE_PATTERN = re.compile(r"\b(\d{1,2})[\s./-](\d{1,2})[\s./-](\d{2,4})\b")


def s(v) -> str:
    return str(v).strip() if v is not None else ""


def is_numeric(x: str) -> bool:
    return bool(re.match(r"^\d+(\.0+)?$", x))


def extract_company_and_date(a_text: str, c_text: str) -> tuple[str, str]:
    """Pull (company, date) from a drive-header row."""
    date = ""
    company = re.sub(r"^company\s*name\s*[-:]\s*", "", a_text, flags=re.IGNORECASE).strip()
    m = re.search(r"date\s*of\s*drive\s*[-:]\s*(.+)$", c_text, re.IGNORECASE)
    if m:
        date = m.group(1).strip()
    elif c_text and DATE_PATTERN.search(c_text):
        date = c_text
    if not date:
        m = DATE_PATTERN.search(company)
        if m:
            date = m.group(0)
    company = DATE_PATTERN.sub("", company).strip()
    company = re.sub(r"\s+", " ", company).strip(" -,.()")
    return company, date


# ---- Canonical recruiter mapping. Order matters — first match wins. ----
# Multi-recruiter pool drives (commas / ampersands) are tagged
# separately so they're counted as drives but don't pollute the
# distinct-company tally or the Top Recruiter slot.
_CANON_MAP: list[tuple[str, str]] = [
    (r"\bHOLLISTER\b", "Hollister"),
    (r"\b(VIKAS|VIKASH)\b.*\bGROUP\b", "Vikas Group"),
    (r"\b(VIKAS|VIKASH)\b", "Vikas Group"),
    (r"\bSADEN\b", "Vikas Group"),
    (r"\bKRISHNA MARUTI\b", "Krishna Maruti"),
    (r"\bMOTHERSON\b", "Motherson Automotives"),
    (r"\bDHOOT\b", "Dhoot Transmission"),
    (r"\bNEW ALLENBERRY\b", "New Allenberry"),
    (r"\bNEW ALEMBERRY\b", "New Allenberry"),
    (r"\bASIAN PAINTS\b", "Asian Paints"),
    (r"\bMAHINDRA\b", "Mahindra & Mahindra"),
    (r"\bJBM\b", "JBM Group"),
    (r"\bR\s*R\s*PARKON\b", "R R Parkon"),
    # Session-2026 drives, added 3 Sep 2026. Without these the parser
    # title-cases the TPO's own spellings into new phantom recruiters:
    # "SHOPOORJI POLONIJI" -> "Shopoorji Poloniji" alongside the real
    # Shapoorji Pallonji, and "KNORR-BREMSE" would never group. Names
    # are taken from the companies' own branding, not the sheet.
    (r"\bKNORR\b", "Knorr-Bremse"),
    (r"\bSH[OA]POORJI\b|\bPALLONJI\b|\bPOLONIJI\b", "Shapoorji Pallonji"),
    (r"\bAMBAR\b", "Ambar Enterprises"),
    (r"\bEATON\b", "Eaton"),
    (r"\bEXIDE\b", "Exide India"),
    (r"\bMINDA\b", "Minda Corporation"),
    (r"\bRADIENTS\b|\bRADIANT\b", "Radients Appliances"),
    (r"\bISMT\b", "ISMT"),
    (r"\bZEMENCRETE\b", "Zemencrete"),
    (r"\bHFCL\b", "HFCL"),
    (r"\bMKC\b", "MKC Gujarat"),
    (r"\bSIDDHESHWAR\b", "Siddheshwar"),
    (r"\bMSKH\b", "MSKH Seating"),
    (r"\bSECURE METERS\b", "Secure Meters"),
    (r"\bTAXTRON\b", "Taxtron Technologies"),
    (r"\bJAIEESPY\b", "Jaieespy Techkart"),
    (r"\bSIGMA\b", "Sigma Tools"),
    (r"\bFORCE FOX\b", "Force Fox"),
    (r"\bBHAGAWTI\b", "Bhagawti Products"),
    (r"\bI CAN NANO\b", "I Can Nano"),
    (r"\bTALBROSS\b|\bTALBRASS\b", "Talbross Group"),
    (r"\bRISHISHWAR\b", "Rishishwar Construction"),
    (r"\bEASHU\b", "Eashu Construction"),
    (r"\bROHANI\b", "Rohani Construction"),
]


def canonicalize_recruiter(name: str) -> str:
    """Map a raw company string to a clean canonical recruiter name."""
    n = DATE_PATTERN.sub("", name.upper()).strip()
    if "," in n or "&" in n:
        return "Multi-recruiter pool drive"
    for pat, target in _CANON_MAP:
        if re.search(pat, n):
            return target
    return name.title()


def normalize_branch(branch_raw: str) -> str:
    """Collapse the dozens of branch spellings in the XLSX to BIPE's
    actual 5 BTEUP branches (plus Mechanical sub-tracks)."""
    if not branch_raw:
        return "Unknown"
    b = branch_raw.upper()
    if "CIVIL" in b or b.strip() == "CE":
        return "Civil Engineering"
    if "ELECTRICAL" in b or b.strip() in ("EE", "ELECTRICAL ENGG", "ELECTRICAL ENGG."):
        return "Electrical Engineering"
    if "ELECTRONICS" in b:
        return "Electronics Engineering"
    if "MECHANICAL" in b and "PRODUCTION" in b:
        return "Mechanical Engineering (Production)"
    if "MECHANICAL" in b and "AUTOMOBILE" in b:
        return "Mechanical Engineering (Automobile)"
    if "MECHANICAL" in b or b.strip() in ("MP", "MA", "ME"):
        return "Mechanical Engineering"
    if "DAIRY" in b:
        return "Dairy Engineering"
    if "COMPUTER" in b or "CSE" in b:
        return "Computer Science & Engineering"
    # Fall-through catches, added 3 Sep 2026. Without these the raw
    # spreadsheet spelling reached the PUBLIC /alumni branch filter as a
    # pill — live examples were "Diploma Electrcal" (sic), "Diploma
    # Automobile Test" and "Production". These re-bucket labels only;
    # no record is added or removed, so the 1,331 total is untouched.
    if "AUTOMOBILE" in b:
        return "Mechanical Engineering (Automobile)"
    if "PRODUCTION" in b:
        return "Mechanical Engineering (Production)"
    if "ELECTRCAL" in b or "ELECTRICL" in b:  # known TPO misspellings
        return "Electrical Engineering"
    return branch_raw.title()


def parse_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    drives_raw: list[dict] = []
    current_drive: dict | None = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        a = s(row[0]) if len(row) > 0 else ""
        b = s(row[1]) if len(row) > 1 else ""
        c = s(row[2]) if len(row) > 2 else ""
        d = s(row[3]) if len(row) > 3 else ""
        if not any([a, b, c, d]):
            continue
        if a.lower() in ("s.no", "s. no", "sno", "sr.no", "sl.no") or b.lower() == "name":
            continue
        # Data row: numeric s.no + name
        if is_numeric(a) and b:
            if current_drive is not None:
                branch_raw = c
                year = None
                # Strip a trailing TPO annotation before looking for the
                # year. The old pattern anchored (\d{4}) at end-of-string,
                # so "CE-2020 (NOT JOIN)" matched nothing: the year was
                # lost and the whole cell became the branch, surfacing on
                # public /alumni as the filter pill "Ce-2020 (Not Join)".
                # The 44 "(NOT JOIN)" records ARE placements — owner
                # confirmed 3 Sep 2026, "1331 is correct, keep it" — so
                # they belong in their real branch and year like any other.
                annotated = re.sub(r"\s*\((?:[^()]*)\)\s*$", "", branch_raw).strip()
                m = re.search(r"(.+?)[-\s.()]*\(?(\d{4})\)?\s*$", annotated or branch_raw)
                if m:
                    branch_raw = m.group(1).rstrip(" -.()")
                    year = m.group(2)
                current_drive["students"].append(
                    {"name": b, "branch_raw": branch_raw, "year": year}
                )
            continue
        # Drive header row: non-empty A, no name in B
        if a and not is_numeric(a) and b == "":
            company, date = extract_company_and_date(a, c)
            if company:
                current_drive = {"company_raw": company, "date": date, "students": []}
                drives_raw.append(current_drive)

    return _build_manifest(drives_raw)


def _build_manifest(drives_raw: list[dict]) -> dict:
    # Normalise companies + branches
    drives: list[dict] = []
    alumni: list[dict] = []

    for dr in drives_raw:
        recruiter = canonicalize_recruiter(dr["company_raw"])
        # Year from drive date if available
        drive_year = None
        m = DATE_PATTERN.search(dr["date"] or "")
        if m:
            y = m.group(3)
            if len(y) == 2:
                y = ("20" if int(y) < 50 else "19") + y
            drive_year = int(y)

        drives.append(
            {
                "company": recruiter,
                "company_raw": dr["company_raw"],
                "date": dr["date"] or "",
                "year": drive_year,
                "joined": len(dr["students"]),
                "offered": 0,
            }
        )

        for st in dr["students"]:
            # Numeric IDs for compat with the backend's getAlumni()
            # rows (which carry int PKs).
            alumni.append(
                {
                    "id": len(alumni) + 1,
                    "name": st["name"],
                    "branch": normalize_branch(st["branch_raw"]),
                    "year": st["year"] or (str(drive_year) if drive_year else None),
                    "company": recruiter,
                    "driveDate": dr["date"] or "",
                    "status": "joined",
                }
            )

    # Compute totals + leaderboards
    single_recruiters = collections.Counter()
    pool_drive_count = 0
    pool_student_count = 0
    for dr in drives:
        if dr["company"] != "Multi-recruiter pool drive":
            single_recruiters[dr["company"]] += dr["joined"]
        else:
            pool_drive_count += 1
            pool_student_count += dr["joined"]

    top_recruiter_name, top_recruiter_count = (
        single_recruiters.most_common(1)[0] if single_recruiters else ("Unknown", 0)
    )

    years = sorted({a["year"] for a in alumni if a["year"]})
    branches = sorted(
        {a["branch"] for a in alumni if a["branch"] and a["branch"] != "Unknown"}
    )

    return {
        # ── Top-level shape preserved for the existing AlumniView reader.
        #    Older fields kept; new fields added.
        "totalAlumni": len(alumni),
        "totalJoined": sum(d["joined"] for d in drives),
        "totalOffered": 0,
        "totalDrives": len(drives),
        # New: distinct single-recruiter companies + pool-drive count
        "distinctRecruiters": len(single_recruiters),
        "poolDriveCount": pool_drive_count,
        "poolStudentCount": pool_student_count,
        "topRecruiter": {"name": top_recruiter_name, "count": top_recruiter_count},
        "recruiterCounts": dict(single_recruiters.most_common()),
        "branches": branches,
        "years": years,
        "drives": drives,
        "alumni": alumni,
    }


def main():
    if not SOURCE_XLSX.exists():
        sys.exit(f"Source file not found: {SOURCE_XLSX}")

    print(f"→ Parsing {SOURCE_XLSX.relative_to(REPO_ROOT)} …")
    manifest = parse_xlsx(SOURCE_XLSX)

    OUTPUT_MANIFEST.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    print()
    print(f"✓ Wrote {OUTPUT_MANIFEST.relative_to(REPO_ROOT)}")
    print()
    print("  Drives                : ", manifest["totalDrives"])
    print("  Named students        : ", manifest["totalAlumni"])
    print("  Distinct recruiters   : ", manifest["distinctRecruiters"])
    print("  Multi-recruiter pools : ", manifest["poolDriveCount"], "drives,",
          manifest["poolStudentCount"], "students")
    print(
        f"  Top recruiter         :  {manifest['topRecruiter']['name']} "
        f"({manifest['topRecruiter']['count']} placements)"
    )
    print("  Years on record       : ", manifest["years"][0], "–", manifest["years"][-1],
          f"({len(manifest['years'])} years)")
    print()
    print("Now run:  npm run build  (the constants in lib/placement-stats.ts")
    print("will pick up the new totals automatically).")


if __name__ == "__main__":
    main()
